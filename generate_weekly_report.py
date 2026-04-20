#!/usr/bin/env python3
"""
Weekly Report Generator
Generates weekly property reports from source Excel files using a template.
Pulls property info and historical data from PostgreSQL database.
"""

import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from datetime import datetime, date, timedelta
import os
import sys
import glob
import time
import re
import zipfile
import shutil
from sqlalchemy import create_engine, Column, String, Integer, Float, Date, UniqueConstraint
from sqlalchemy.orm import sessionmaker, declarative_base
from sqlalchemy.exc import OperationalError

def _fix_xlsx_rels(filepath, template_path=None):
    """
    Fix openpyxl xlsx output by copying .rels files from the original template.
    openpyxl rewrites relationship files with absolute paths and phantom references
    that Excel can't handle. The safest fix is to restore the original .rels files.
    """
    if template_path is None:
        for candidate in [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'Weekly_Report_Template_Clean.xlsx'),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'Long_Report_Template.xlsx'),
        ]:
            if os.path.exists(candidate):
                template_path = candidate
                break
    if not template_path or not os.path.exists(template_path):
        return

    temp_path = filepath + '.tmp'
    with zipfile.ZipFile(template_path, 'r') as zt, \
         zipfile.ZipFile(filepath, 'r') as zin, \
         zipfile.ZipFile(temp_path, 'w') as zout:
        template_rels = {n: zt.read(n) for n in zt.namelist() if n.endswith('.rels')}
        for item in zin.infolist():
            if item.filename.endswith('.rels') and item.filename in template_rels:
                zout.writestr(item, template_rels[item.filename])
            else:
                zout.writestr(item, zin.read(item.filename))
    shutil.move(temp_path, filepath)


# Database connection
DATABASE_URL = "postgresql://postgres:ysBxQxKGOxlvIFfhCVrEYlApmrqElAMB@shinkansen.proxy.rlwy.net:55881/railway"
DB_CONNECT_RETRIES = 3
DB_RETRY_DELAY = 5  # seconds

Base = declarative_base()

class Property(Base):
    __tablename__ = "properties"
    property_code = Column(String(20), primary_key=True)
    name = Column(String(100), nullable=False)
    units = Column(Integer, nullable=True)
    models = Column(String(200), nullable=True)
    office = Column(String(100), nullable=True)
    location = Column(String(200), nullable=True)
    equity_investment = Column(Float, nullable=True)
    analysis_start = Column(Date, nullable=True)

class HistoricalOccupancy(Base):
    __tablename__ = "historical_occupancy"
    id = Column(Integer, primary_key=True, autoincrement=True)
    property_code = Column(String(20), nullable=False)
    date = Column(Date, nullable=False)
    occupancy = Column(Float, nullable=True)
    leased = Column(Float, nullable=True)
    projection = Column(Float, nullable=True)
    make_ready = Column(Integer, nullable=True)
    work_orders = Column(Integer, nullable=True)
    __table_args__ = (
        UniqueConstraint("property_code", "date", name="uix_historical_occupancy"),
    )

class HistoricalFinancial(Base):
    __tablename__ = "historical_financial"
    id = Column(Integer, primary_key=True, autoincrement=True)
    property_code = Column(String(20), nullable=False)
    date = Column(Date, nullable=False)
    market_rent = Column(Float, nullable=True)
    occupied_rent = Column(Float, nullable=True)
    revenue = Column(Float, nullable=True)
    expenses = Column(Float, nullable=True)
    owed = Column(Float, nullable=True)
    charges = Column(Float, nullable=True)
    collections = Column(Float, nullable=True)
    collections_pct = Column(Float, nullable=True)
    __table_args__ = (
        UniqueConstraint("property_code", "date", name="uix_historical_financial"),
    )


def get_db_session():
    """Get database session with retry logic. Returns None if DB is unreachable."""
    from sqlalchemy import text
    for attempt in range(1, DB_CONNECT_RETRIES + 1):
        try:
            engine = create_engine(DATABASE_URL, echo=False, connect_args={"connect_timeout": 10})
            Session = sessionmaker(bind=engine)
            session = Session()
            session.execute(text("SELECT 1"))  # Test connection
            return session
        except Exception as e:
            if attempt < DB_CONNECT_RETRIES:
                print(f"  ⚠ DB connection attempt {attempt}/{DB_CONNECT_RETRIES} failed: {type(e).__name__}")
                print(f"    Retrying in {DB_RETRY_DELAY}s...")
                time.sleep(DB_RETRY_DELAY)
            else:
                print(f"  ✗ DB unreachable after {DB_CONNECT_RETRIES} attempts: {type(e).__name__}")
                print(f"    Will generate report WITHOUT historical data.")
                return None


def get_property_info(session, property_code):
    """Get property info from database."""
    prop = session.query(Property).filter_by(property_code=property_code).first()
    if prop:
        return {
            'name': prop.name,
            'units': prop.units,
            'models': prop.models,
            'office': prop.office,
            'location': prop.location,
            'equity_investment': prop.equity_investment,
            'analysis_start': prop.analysis_start
        }
    return None


def get_historical_occupancy(session, property_code):
    """Get all historical occupancy data from database, ordered by date."""
    records = session.query(HistoricalOccupancy).filter_by(
        property_code=property_code
    ).order_by(HistoricalOccupancy.date).all()

    return records


def get_historical_financial(session, property_code):
    """Get all historical financial data from database, ordered by date."""
    records = session.query(HistoricalFinancial).filter_by(
        property_code=property_code
    ).order_by(HistoricalFinancial.date).all()

    return records


def get_weekly_data(engine, property_code, week_date):
    """Get weekly leasing/maintenance data for a specific week."""
    from sqlalchemy import text
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT new_leads, tours, applications, move_ins, move_outs,
                   make_ready_count, work_orders_count
            FROM weekly_data
            WHERE property_code = :prop AND week_date = :wdate
        """), {"prop": property_code, "wdate": week_date})
        row = result.fetchone()
        if row:
            return {
                'new_leads': row[0] or 0,
                'tours': row[1] or 0,
                'applications': row[2] or 0,
                'move_ins': row[3] or 0,
                'move_outs': row[4] or 0,
                'make_ready': row[5] or 0,
                'work_orders': row[6] or 0
            }
    return None


def get_monthly_leasing_totals(engine, property_code, year, month):
    """Sum weekly leasing data for a given month."""
    from sqlalchemy import text
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT SUM(new_leads), SUM(tours), SUM(applications),
                   SUM(move_ins), SUM(move_outs)
            FROM weekly_data
            WHERE property_code = :prop
              AND EXTRACT(YEAR FROM week_date) = :yr
              AND EXTRACT(MONTH FROM week_date) = :mo
        """), {"prop": property_code, "yr": year, "mo": month})
        row = result.fetchone()
        if row and row[0] is not None:
            return {
                'new_leads': int(row[0] or 0),
                'tours': int(row[1] or 0),
                'applications': int(row[2] or 0),
                'move_ins': int(row[3] or 0),
                'move_outs': int(row[4] or 0)
            }
    return None


def get_last_month_data(engine, property_code, year, month):
    """Get monthly data from monthly_data table."""
    from sqlalchemy import text
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT new_leads, tours, applications, move_ins, move_outs
            FROM monthly_data
            WHERE property_code = :prop
              AND EXTRACT(YEAR FROM month_date) = :yr
              AND EXTRACT(MONTH FROM month_date) = :mo
        """), {"prop": property_code, "yr": year, "mo": month})
        row = result.fetchone()
        if row:
            return {
                'new_leads': row[0] or 0,
                'tours': row[1] or 0,
                'applications': row[2] or 0,
                'move_ins': row[3] or 0,
                'move_outs': row[4] or 0
            }
    return None


def save_current_week_data(engine, property_code, week_date, new_leads, tours, applications,
                           move_ins, move_outs, make_ready_count, work_orders_count):
    """Save or update current week's leasing/maintenance data to weekly_data table."""
    from sqlalchemy import text
    with engine.connect() as conn:
        # Check if record exists
        result = conn.execute(text("""
            SELECT id FROM weekly_data
            WHERE property_code = :prop AND week_date = :wdate
        """), {"prop": property_code, "wdate": week_date})
        existing = result.fetchone()

        if existing:
            conn.execute(text("""
                UPDATE weekly_data
                SET new_leads = :leads, tours = :tours, applications = :apps,
                    move_ins = :mi, move_outs = :mo,
                    make_ready_count = :mr, work_orders_count = :wo
                WHERE property_code = :prop AND week_date = :wdate
            """), {
                "prop": property_code, "wdate": week_date,
                "leads": new_leads, "tours": tours, "apps": applications,
                "mi": move_ins, "mo": move_outs, "mr": make_ready_count, "wo": work_orders_count
            })
        else:
            conn.execute(text("""
                INSERT INTO weekly_data (property_code, week_date, new_leads, tours, applications,
                                         move_ins, move_outs, make_ready_count, work_orders_count)
                VALUES (:prop, :wdate, :leads, :tours, :apps, :mi, :mo, :mr, :wo)
            """), {
                "prop": property_code, "wdate": week_date,
                "leads": new_leads, "tours": tours, "apps": applications,
                "mi": move_ins, "mo": move_outs, "mr": make_ready_count, "wo": work_orders_count
            })
        conn.commit()


def save_current_week_occupancy(session, property_code, report_date, occupancy_pct, leased_pct, projection_pct, make_ready, work_orders):
    """Save or update current week's occupancy data to database."""
    existing = session.query(HistoricalOccupancy).filter_by(
        property_code=property_code,
        date=report_date
    ).first()

    if existing:
        existing.occupancy = occupancy_pct
        existing.leased = leased_pct
        existing.projection = projection_pct
        existing.make_ready = make_ready
        existing.work_orders = work_orders
    else:
        new_record = HistoricalOccupancy(
            property_code=property_code,
            date=report_date,
            occupancy=occupancy_pct,
            leased=leased_pct,
            projection=projection_pct,
            make_ready=make_ready,
            work_orders=work_orders
        )
        session.add(new_record)

    session.commit()


def save_current_week_financial(session, property_code, report_date, market_rent, occupied_rent, revenue, expenses, owed, charges, collections):
    """Save or update current week's financial data to database."""
    existing = session.query(HistoricalFinancial).filter_by(
        property_code=property_code,
        date=report_date
    ).first()

    if existing:
        existing.market_rent = market_rent
        existing.occupied_rent = occupied_rent
        existing.revenue = revenue
        existing.expenses = expenses
        existing.owed = owed
        existing.charges = charges
        existing.collections = collections
    else:
        new_record = HistoricalFinancial(
            property_code=property_code,
            date=report_date,
            market_rent=market_rent,
            occupied_rent=occupied_rent,
            revenue=revenue,
            expenses=expenses,
            owed=owed,
            charges=charges,
            collections=collections
        )
        session.add(new_record)

    session.commit()


def find_file(directory, pattern):
    """Find a file matching the pattern in the directory, excluding temp files."""
    matches = [f for f in glob.glob(os.path.join(directory, pattern))
               if not os.path.basename(f).startswith('~$')]
    if matches:
        return matches[0]
    return None


def extract_property_code(directory):
    """Extract property code from directory or file names."""
    import re

    # Known property codes from config
    known_codes = [
        "55pharr", "abbeylk", "capella2", "colwds", "emersn", "georget",
        "hampec", "hangar", "haven", "kenplc", "longvw", "manwes",
        "marbla", "marshp", "perryh", "portico", "talloak", "tapeprk",
        "turn", "wdlndcm"
    ]

    # Try to find property code from file names
    files = glob.glob(os.path.join(directory, "*.xlsx"))
    for f in files:
        basename = os.path.basename(f).lower()
        # Check for known property codes in filename
        for code in known_codes:
            if f"_{code}." in basename or f"_{code}_" in basename or basename.endswith(f"_{code}.xlsx"):
                return code

    # Try to extract from filename pattern like _perryh.xlsx
    for f in files:
        basename = os.path.basename(f).lower().replace('.xlsx', '')
        # Match pattern: ending with _<code> where code is alphanumeric
        match = re.search(r'_([a-z0-9]+)$', basename)
        if match:
            code = match.group(1)
            if code not in ['2', 'accrual', 'details', 'report', 'summary']:
                return code

    # Fallback: use directory name mapped to property code
    dir_name = os.path.basename(directory).lower().replace(' ', '')
    dir_to_code = {
        'perryheights': 'perryh',
        '55pharr': '55pharr',
        'abbeylake': 'abbeylk',
        'capella': 'capella2',
        'colonywoods': 'colwds',
        'emerson1600': 'emersn',
        'georgetown': 'georget',
        'hamptonsateastcobb': 'hampec',
        'thehangar': 'hangar',
        'haven': 'haven',
        'kensingtonplace': 'kenplc',
        'longview': 'longvw',
        'manchesteratweslyn': 'manwes',
        'marbella': 'marbla',
        'marshpoint': 'marshp',
        'porticoatlanier': 'portico',
        'talloaks': 'talloak',
        'tapestrypark': 'tapeprk',
        'theturn': 'turn',
        'woodlandcommons': 'wdlndcm'
    }

    return dir_to_code.get(dir_name, dir_name[:10])


def extract_box_score_data(filepath):
    """Extract occupancy and unit data from Box Score Summary."""
    df = pd.read_excel(filepath)

    data = {
        'units': 0,
        'occupied_no_notice': 0,
        'vacant_rented': 0,
        'vacant_unrented': 0,
        'notice_rented': 0,
        'notice_unrented': 0,
        'available': 0,
        'pct_occupied': 0,
        'pct_leased': 0,
        'unit_types': []
    }

    for idx, row in df.iterrows():
        if pd.notna(row.iloc[0]) and str(row.iloc[0]).startswith('1'):
            unit_type = {
                'code': str(row.iloc[0]),
                'name': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
                'units': int(row.iloc[4]) if pd.notna(row.iloc[4]) else 0,
                'vacant_rented': int(row.iloc[6]) if pd.notna(row.iloc[6]) else 0,
                'vacant_unrented': int(row.iloc[7]) if pd.notna(row.iloc[7]) else 0,
                'notice_rented': int(row.iloc[8]) if pd.notna(row.iloc[8]) else 0,
                'notice_unrented': int(row.iloc[9]) if pd.notna(row.iloc[9]) else 0,
                'available': int(row.iloc[10]) if pd.notna(row.iloc[10]) else 0,
            }
            data['unit_types'].append(unit_type)

        if pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() == 'Total':
            data['units'] = int(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
            data['occupied_no_notice'] = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
            data['vacant_rented'] = int(row.iloc[6]) if pd.notna(row.iloc[6]) else 0
            data['vacant_unrented'] = int(row.iloc[7]) if pd.notna(row.iloc[7]) else 0
            data['notice_rented'] = int(row.iloc[8]) if pd.notna(row.iloc[8]) else 0
            data['notice_unrented'] = int(row.iloc[9]) if pd.notna(row.iloc[9]) else 0
            data['available'] = int(row.iloc[10]) if pd.notna(row.iloc[10]) else 0

            pct_occ = row.iloc[15] if len(row) > 15 and pd.notna(row.iloc[15]) else 0
            pct_leased = row.iloc[16] if len(row) > 16 and pd.notna(row.iloc[16]) else 0
            data['pct_occupied'] = float(pct_occ) / 100 if pct_occ else 0
            data['pct_leased'] = float(pct_leased) / 100 if pct_leased else 0
            break

    return data


def extract_lease_expiration_data(filepath):
    """Extract lease expiration data."""
    df = pd.read_excel(filepath)

    months = df.iloc[2, 3:16].tolist()
    values = df.iloc[3, 3:16].tolist()
    values = [int(v) if pd.notna(v) else 0 for v in values]

    return {
        'mtm': values[0],
        'months': months[1:],
        'values': values[1:]
    }


def extract_delinquency_data(filepath):
    """Extract delinquency summary data."""
    df = pd.read_excel(filepath)

    data = {
        'delinq_0_30': 0,
        'delinq_31_60': 0,
        'delinq_61_90': 0,
        'delinq_over_90': 0,
        'prepayments': 0,
        'total_owed': 0
    }

    for idx, row in df.iterrows():
        row_str = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        if 'Total' in row_str and 'Grand' not in row_str:
            data['delinq_0_30'] = float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0
            data['delinq_31_60'] = float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
            data['delinq_61_90'] = float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
            data['delinq_over_90'] = float(row.iloc[6]) if pd.notna(row.iloc[6]) else 0
            data['prepayments'] = float(row.iloc[7]) if pd.notna(row.iloc[7]) else 0
            data['total_owed'] = float(row.iloc[8]) if pd.notna(row.iloc[8]) else 0
            break

    return data


def extract_conversion_data(filepath):
    """Extract leasing/conversion data."""
    df = pd.read_excel(filepath)

    data = {'new_leads': 0, 'tours': 0, 'applications': 0}

    for idx, row in df.iterrows():
        row_str = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        # Skip header rows and look for property code row (has numeric data in columns)
        # The data row has the property code without parentheses and has numeric values
        if row_str and not row_str.startswith('Conversion') and not row_str.startswith('Date') and not row_str.startswith('Property'):
            if 'Total' not in row_str and 'Weighted' not in row_str and '(' not in row_str:
                # Verify this row has numeric data (not NaN)
                if pd.notna(row.iloc[3]):
                    calls = int(row.iloc[3]) if pd.notna(row.iloc[3]) else 0
                    walkin = int(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
                    email = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
                    web = int(row.iloc[8]) if pd.notna(row.iloc[8]) else 0
                    data['new_leads'] = calls + walkin + email + web
                    data['tours'] = int(row.iloc[11]) if pd.notna(row.iloc[11]) else 0
                    data['applications'] = int(row.iloc[13]) if pd.notna(row.iloc[13]) else 0
                    break

    return data


def extract_residents_on_notice(filepath):
    """Extract eviction count from Residents on Notice file."""
    df = pd.read_excel(filepath)

    evictions_filed = 0

    for idx, row in df.iterrows():
        status = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''
        if status == 'Eviction':
            evictions_filed += 1

    return {'evictions_filed': evictions_filed}


def extract_not_filed_count(filepath):
    """Extract 'not filed' count from detailed delinquency report.

    Counts residents where status is 'Current' and 0-30 owed > $1000.
    """
    df = pd.read_excel(filepath)

    not_filed = 0

    for idx, row in df.iterrows():
        status = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
        # Skip header rows - only process rows where column 6 is numeric
        try:
            owed_0_30 = float(row.iloc[6]) if pd.notna(row.iloc[6]) else 0
        except (ValueError, TypeError):
            continue

        if status == 'Current' and owed_0_30 > 1000:
            not_filed += 1

    return not_filed


def extract_make_ready_count(filepath):
    """Count pending make ready units."""
    df = pd.read_excel(filepath)
    count = 0
    for idx, row in df.iterrows():
        row_str = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        if row_str and row_str not in ['pending', 'property', 'code', '', 'nan'] and 'property=' not in row_str:
            count += 1
    return count


def extract_work_order_count(filepath):
    """Count open work orders."""
    df = pd.read_excel(filepath)
    count = 0
    for idx, row in df.iterrows():
        row_str = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        if row_str.isdigit():
            count += 1
    return count


def extract_resident_activity(filepath):
    """Extract move-ins and move-outs from Box Score (current week)."""
    df = pd.read_excel(filepath)

    move_ins = 0
    move_outs = 0

    in_resident_activity = False
    for idx, row in df.iterrows():
        row_str = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        if 'Resident Activity' in row_str:
            in_resident_activity = True
        if in_resident_activity and pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() == 'Total':
            move_ins = int(row.iloc[3]) if pd.notna(row.iloc[3]) else 0
            move_outs = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
            break

    return {'move_ins': move_ins, 'move_outs': move_outs}


def extract_projected_occupancy(filepath):
    """Extract 30-day (weeks 1-4) move-ins and move-outs from Projected Occupancy report."""
    df = pd.read_excel(filepath)

    move_ins_30day = 0
    move_outs_30day = 0

    # Find the data rows (after the header rows, starting with date pattern)
    data_rows = []
    for idx, row in df.iterrows():
        row_str = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        # Look for date patterns like "02/09/2026"
        if '/' in row_str and len(row_str) == 10:
            data_rows.append(idx)

    # Sum weeks 1-4 (first 4 data rows)
    for i, idx in enumerate(data_rows[:4]):
        row = df.iloc[idx]
        move_ins = int(row.iloc[2]) if pd.notna(row.iloc[2]) else 0
        move_outs = int(row.iloc[3]) if pd.notna(row.iloc[3]) else 0
        move_ins_30day += move_ins
        move_outs_30day += move_outs

    return {'move_ins_30day': move_ins_30day, 'move_outs_30day': move_outs_30day}


def generate_report(data_dir, template_path, output_path, report_date=None):
    """Generate a weekly report from source data and database."""

    print(f"Generating report from: {data_dir}")

    # Get database session (may be None if DB is unreachable)
    session = get_db_session()
    db_available = session is not None

    # Extract property code from directory/files
    property_code = extract_property_code(data_dir)
    print(f"  Property code: {property_code}")

    # Get property info from database
    property_info = get_property_info(session, property_code) if db_available else None
    if property_info:
        print(f"  Property name: {property_info['name']}")
    else:
        if db_available:
            print(f"  Warning: No property info found in database for '{property_code}'")
        property_info = {'name': property_code, 'units': 0, 'models': '', 'office': '',
                        'location': '', 'equity_investment': 0, 'analysis_start': None}

    # Find source files
    box_score_file = find_file(data_dir, "*Box_Score*.xlsx")
    lease_exp_file = find_file(data_dir, "*Lease_Expiration*.xlsx")
    # Get delinquency files - summary for totals, detail (_2) for not filed count
    # Exclude Excel temp files (~$)
    delinquency_matches = [f for f in glob.glob(os.path.join(data_dir, "*Delinquency_Summary*.xlsx")) if not os.path.basename(f).startswith('~$')]
    delinquency_file = next((f for f in delinquency_matches if '_2.' not in f), delinquency_matches[0] if delinquency_matches else None)
    delinquency_detail_file = next((f for f in delinquency_matches if '_2.' in f), None)
    conversion_file = find_file(data_dir, "*Conversion_Ratios*.xlsx")
    notice_file = find_file(data_dir, "*Residents_on_Notice*.xlsx")
    make_ready_file = find_file(data_dir, "*Make_Ready*.xlsx")
    work_order_file = find_file(data_dir, "*Work_Order*.xlsx")
    projected_occ_file = find_file(data_dir, "*Projected*Occupancy*.xlsx")
    if not projected_occ_file:
        projected_occ_file = find_file(data_dir, "*ProjectedOccupancy*.xlsx")

    # Extract data from each file
    print("  Extracting data from source files...")
    box_score = extract_box_score_data(box_score_file) if box_score_file else {}
    activity = extract_resident_activity(box_score_file) if box_score_file else {}  # Current week move ins/outs
    projected_occ = extract_projected_occupancy(projected_occ_file) if projected_occ_file else {}  # 30-day move ins/outs
    lease_exp = extract_lease_expiration_data(lease_exp_file) if lease_exp_file else {}
    delinquency = extract_delinquency_data(delinquency_file) if delinquency_file else {}
    conversion = extract_conversion_data(conversion_file) if conversion_file else {}
    notices = extract_residents_on_notice(notice_file) if notice_file else {}
    not_filed_count = extract_not_filed_count(delinquency_detail_file) if delinquency_detail_file else 0
    make_ready_count = extract_make_ready_count(make_ready_file) if make_ready_file else 0
    work_order_count = extract_work_order_count(work_order_file) if work_order_file else 0

    # Get historical data from database
    if db_available:
        print("  Loading historical data from database...")
        historical_occupancy = get_historical_occupancy(session, property_code)
        historical_financial = get_historical_financial(session, property_code)
        print(f"  Found {len(historical_occupancy)} historical occupancy records")
        print(f"  Found {len(historical_financial)} historical financial records")
    else:
        print("  ⚠ Skipping historical data (DB unavailable)")
        historical_occupancy = []
        historical_financial = []

    # Determine report date (default to Monday of current week)
    if report_date is None:
        today = date.today()
        # Calculate Monday of current week (weekday() returns 0 for Monday)
        report_date = today - timedelta(days=today.weekday())
    elif isinstance(report_date, str):
        report_date = datetime.strptime(report_date, '%Y-%m-%d').date()

    # Current week's data for historical tracking
    current_occupancy_pct = box_score.get('pct_occupied', 0) * 100
    current_leased_pct = box_score.get('pct_leased', 0) * 100
    current_projection_pct = current_occupancy_pct  # Usually same as occupancy

    # Save current week to database
    if db_available:
        print("  Saving current week's data to database...")
        save_current_week_occupancy(
            session, property_code, report_date,
            current_occupancy_pct, current_leased_pct, current_projection_pct,
            make_ready_count, work_order_count
        )
    else:
        print("  ⚠ Skipping DB save (DB unavailable)")

    # Save current week's leasing/maintenance data to weekly_data table (only if not already saved)
    engine = create_engine(DATABASE_URL, echo=False) if db_available else None
    existing_week = get_weekly_data(engine, property_code, report_date) if db_available else None
    if db_available and not existing_week:
        save_current_week_data(
            engine, property_code, report_date,
            conversion.get('new_leads', 0),
            conversion.get('tours', 0),
            conversion.get('applications', 0),
            activity.get('move_ins', 0),
            activity.get('move_outs', 0),
            make_ready_count,
            work_order_count
        )
        print(f"  Saved weekly data for {report_date}")
    elif db_available:
        print(f"  Weekly data for {report_date} already exists, skipping")

    # Load template
    print("  Loading template and populating...")
    wb = load_workbook(template_path)

    # === OCCUPANCY SHEET ===
    ws = wb['Occupancy']

    # Property info from database (Row 2-8)
    ws['B2'] = property_info['name']
    ws['B3'] = property_info['units'] or box_score.get('units', 0)
    ws['B4'] = property_info['models'] or ''
    ws['B5'] = property_info['office'] or ''
    ws['B6'] = property_info['analysis_start'] if property_info['analysis_start'] else ''
    ws['B7'] = property_info['location'] or ''
    ws['B8'] = property_info['equity_investment'] or ''

    # Report date (F1 is used by formula in L2 for "Weekly Report | date")
    ws['F1'] = report_date
    ws['C3'] = report_date

    # Current Occupancy VALUES in Row 17
    ws['E17'] = box_score.get('pct_occupied', 0)
    ws['F17'] = box_score.get('pct_leased', 0)
    ws['H17'] = box_score.get('units', 0)
    # Current Occupied = Occupied No Notice + Notice Rented + Notice Unrented
    ws['I17'] = (box_score.get('occupied_no_notice', 0) +
                 box_score.get('notice_rented', 0) +
                 box_score.get('notice_unrented', 0))
    ws['J17'] = box_score.get('vacant_rented', 0) + box_score.get('vacant_unrented', 0)
    ws['K17'] = box_score.get('notice_rented', 0) + box_score.get('notice_unrented', 0)
    ws['L17'] = box_score.get('vacant_rented', 0) + box_score.get('notice_rented', 0)

    # 30 Day Projected VALUES in Row 19
    # Use sum of weeks 1-4 from Projected Occupancy report for 30-day move ins/outs
    move_ins_30day = projected_occ.get('move_ins_30day', 0) if projected_occ else activity.get('move_ins', 0)
    move_outs_30day = projected_occ.get('move_outs_30day', 0) if projected_occ else activity.get('move_outs', 0)

    # Calculate projected occupancy: (Current Occupied + Move Ins - Move Outs) / Total Units
    # Current Occupied = Occupied No Notice + Notice Rented + Notice Unrented
    current_occupied = (box_score.get('occupied_no_notice', 0) +
                        box_score.get('notice_rented', 0) +
                        box_score.get('notice_unrented', 0))
    total_units = box_score.get('units', 1)  # Avoid division by zero
    projected_occupied = current_occupied + move_ins_30day - move_outs_30day
    projected_occupancy_pct = projected_occupied / total_units if total_units > 0 else 0

    ws['E19'] = projected_occupancy_pct
    ws['F19'] = box_score.get('available', 0)
    ws['G19'] = move_ins_30day
    ws['I19'] = move_outs_30day
    ws['J19'] = notices.get('evictions_filed', 0)

    # Unit Vacancy Detail - Row 22+
    row = 22
    for ut in box_score.get('unit_types', []):
        ws[f'E{row}'] = ut['code']
        ws[f'F{row}'] = ut['name']
        ws[f'G{row}'] = ut['units']
        ws[f'H{row}'] = ut['vacant_rented']
        ws[f'I{row}'] = ut['vacant_unrented']
        ws[f'J{row}'] = ut['notice_rented']
        ws[f'K{row}'] = ut['notice_unrented']
        ws[f'L{row}'] = ut['available']
        row += 1

    # Historical Occupancy Data - Columns N-S, starting at row 22
    # Headers are in row 21: Date, Occupancy, Leased, Projection, Make Ready, Work Orders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    pct_format = '0.00%'
    date_format = 'm/d/yy'

    # Copy fill style from existing template row 22
    template_fills = {}
    for col in ['N', 'O', 'P', 'Q', 'R', 'S']:
        template_fills[col] = copy(ws[f'{col}22'].fill)

    hist_row = 22
    for record in historical_occupancy:
        ws[f'N{hist_row}'] = record.date
        ws[f'N{hist_row}'].number_format = date_format
        ws[f'N{hist_row}'].border = thin_border
        ws[f'N{hist_row}'].fill = template_fills['N']

        ws[f'O{hist_row}'] = record.occupancy / 100 if record.occupancy else None
        ws[f'O{hist_row}'].number_format = pct_format
        ws[f'O{hist_row}'].border = thin_border
        ws[f'O{hist_row}'].fill = template_fills['O']

        ws[f'P{hist_row}'] = record.leased / 100 if record.leased else None
        ws[f'P{hist_row}'].number_format = pct_format
        ws[f'P{hist_row}'].border = thin_border
        ws[f'P{hist_row}'].fill = template_fills['P']

        ws[f'Q{hist_row}'] = record.projection / 100 if record.projection else None
        ws[f'Q{hist_row}'].number_format = pct_format
        ws[f'Q{hist_row}'].border = thin_border
        ws[f'Q{hist_row}'].fill = template_fills['Q']

        ws[f'R{hist_row}'] = record.make_ready
        ws[f'R{hist_row}'].border = thin_border
        ws[f'R{hist_row}'].fill = template_fills['R']

        ws[f'S{hist_row}'] = record.work_orders
        ws[f'S{hist_row}'].border = thin_border
        ws[f'S{hist_row}'].fill = template_fills['S']

        hist_row += 1

    print(f"  Populated {len(historical_occupancy)} historical occupancy rows with borders and fill")

    # Style the Occupancy chart
    if len(ws._charts) > 0 and len(historical_occupancy) > 0:
        from openpyxl.chart.label import DataLabelList

        occ_chart = ws._charts[0]

        # Make axes visible
        occ_chart.x_axis.delete = False
        occ_chart.y_axis.delete = False

        # Set axis tick label positions
        occ_chart.x_axis.tickLblPos = 'low'
        occ_chart.y_axis.tickLblPos = 'low'

        # Set y-axis range (75% to 100%) and format
        occ_chart.y_axis.scaling.min = 0.75
        occ_chart.y_axis.scaling.max = 1.0
        occ_chart.y_axis.number_format = '0%'

        # Set x-axis date format
        occ_chart.x_axis.number_format = 'mmm-yy'

        # Update data ranges to show only last 13 months (approx 56 weeks)
        end_row = 21 + len(historical_occupancy)
        # Calculate start row for last 13 months of weekly data (56 weeks)
        weeks_to_show = 56
        if len(historical_occupancy) > weeks_to_show:
            chart_start_row = end_row - weeks_to_show + 1
        else:
            chart_start_row = 22

        for series in occ_chart.series:
            if series.val and series.val.numRef:
                ref = series.val.numRef.f
                if '$O$' in ref:
                    series.val.numRef.f = f"Occupancy!$O${chart_start_row}:$O${end_row}"
                elif '$P$' in ref:
                    series.val.numRef.f = f"Occupancy!$P${chart_start_row}:$P${end_row}"
                elif '$Q$' in ref:
                    series.val.numRef.f = f"Occupancy!$Q${chart_start_row}:$Q${end_row}"

            # Update categories (dates)
            if series.cat and series.cat.numRef:
                series.cat.numRef.f = f"Occupancy!$N${chart_start_row}:$N${end_row}"

            # Remove ALL labels
            labels = DataLabelList()
            labels.showVal = False
            labels.showCatName = False
            labels.showSerName = False
            labels.showPercent = False
            labels.showLegendKey = False
            series.labels = labels

            # Set distinct colors for each series
            if series.val and series.val.numRef:
                ref = series.val.numRef.f
                if '$O$' in ref:  # Occupancy - Dark blue
                    series.graphicalProperties.solidFill = "1F4E79"
                elif '$P$' in ref:  # Leased - Medium blue
                    series.graphicalProperties.solidFill = "2E75B6"
                elif '$Q$' in ref:  # Projected - Light blue
                    series.graphicalProperties.solidFill = "9DC3E6"

        print(f"  Styled occupancy chart (y-axis: 0%-100%, data rows: 22-{end_row})")

    # === FINANCIAL SHEET ===
    ws = wb['Financial']

    # Lease Expirations - G16-G27 for months, H16-H27 for counts
    months = lease_exp.get('months', [])
    values = lease_exp.get('values', [])
    for i, (m, v) in enumerate(zip(months, values)):
        if i < 12:
            ws[f'G{16+i}'] = m
            ws[f'H{16+i}'] = v

    # Leasing section header - set current month name
    ws['E28'] = report_date.strftime('%B').upper()  # e.g., "FEBRUARY"

    # Get database engine for weekly/monthly queries
    if not db_available:
        engine = None

    # Calculate date ranges
    last_week_date = report_date - timedelta(days=7)
    two_weeks_ago = report_date - timedelta(days=14)
    three_weeks_ago = report_date - timedelta(days=21)
    current_month = report_date.month
    current_year = report_date.year
    last_month = current_month - 1 if current_month > 1 else 12
    last_month_year = current_year if current_month > 1 else current_year - 1

    # Get historical weekly data
    last_week_data = get_weekly_data(engine, property_code, last_week_date) if db_available else None
    two_weeks_data = get_weekly_data(engine, property_code, two_weeks_ago) if db_available else None
    three_weeks_data = get_weekly_data(engine, property_code, three_weeks_ago) if db_available else None

    # Get monthly totals
    current_month_totals = get_monthly_leasing_totals(engine, property_code, current_year, current_month) if db_available else None
    last_month_data = get_last_month_data(engine, property_code, last_month_year, last_month) if db_available else None

    # Leasing section - THIS WEEK (Column C)
    ws['C29'] = conversion.get('new_leads', 0)
    ws['C30'] = conversion.get('tours', 0)
    ws['C31'] = conversion.get('applications', 0)
    ws['C32'] = activity.get('move_ins', 0)
    ws['C33'] = activity.get('move_outs', 0)

    # Leasing section - LAST WEEK (Column D)
    if last_week_data:
        ws['D29'] = last_week_data['new_leads']
        ws['D30'] = last_week_data['tours']
        ws['D31'] = last_week_data['applications']
        ws['D32'] = last_week_data['move_ins']
        ws['D33'] = last_week_data['move_outs']

    # Leasing section - CURRENT MONTH (Column E)
    if current_month_totals:
        ws['E29'] = current_month_totals['new_leads']
        ws['E30'] = current_month_totals['tours']
        ws['E31'] = current_month_totals['applications']
        ws['E32'] = current_month_totals['move_ins']
        ws['E33'] = current_month_totals['move_outs']

    # Leasing section - LAST MONTH (Column F)
    if last_month_data:
        ws['F29'] = last_month_data['new_leads']
        ws['F30'] = last_month_data['tours']
        ws['F31'] = last_month_data['applications']
        ws['F32'] = last_month_data['move_ins']
        ws['F33'] = last_month_data['move_outs']

    # Delinquency section
    ws['H35'] = delinquency.get('delinq_0_30', 0)
    ws['H36'] = delinquency.get('delinq_31_60', 0)
    ws['H37'] = delinquency.get('delinq_61_90', 0)
    ws['H38'] = delinquency.get('delinq_over_90', 0)
    ws['H39'] = delinquency.get('prepayments', 0)
    ws['H40'] = delinquency.get('total_owed', 0)

    # Maintenance section - THIS WEEK (Column C)
    ws['C46'] = make_ready_count
    ws['C47'] = work_order_count

    # Maintenance section - LAST WEEK (Column D)
    if last_week_data:
        ws['D46'] = last_week_data['make_ready']
        ws['D47'] = last_week_data['work_orders']

    # Maintenance section - 2 WEEKS AGO (Column E)
    if two_weeks_data:
        ws['E46'] = two_weeks_data['make_ready']
        ws['E47'] = two_weeks_data['work_orders']

    # Maintenance section - 3 WEEKS AGO (Column F)
    if three_weeks_data:
        ws['F46'] = three_weeks_data['make_ready']
        ws['F47'] = three_weeks_data['work_orders']

    # Evictions section
    ws['H46'] = notices.get('evictions_filed', 0)
    ws['H47'] = not_filed_count  # Current residents with 0-30 owed > $1000
    ws['H48'] = lease_exp.get('mtm', 0)

    # Historical Financial Data - Columns M-T, starting at row 3 (headers in row 2)
    # Headers: Date, Market Rent, Occupied Rent, Revenue, Expenses, Owed, Charges, Collections
    currency_format = '$#,##0.00'
    pct_format_fin = '0.00%'
    date_format_fin = 'm/d/yy'

    # Copy fill style from existing template row 3
    fin_template_fills = {}
    for col in ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
        fin_template_fills[col] = copy(ws[f'{col}3'].fill)

    fin_row = 3
    for record in historical_financial:
        ws[f'M{fin_row}'] = record.date
        ws[f'M{fin_row}'].number_format = date_format_fin
        ws[f'M{fin_row}'].border = thin_border
        ws[f'M{fin_row}'].fill = fin_template_fills['M']

        ws[f'N{fin_row}'] = record.market_rent
        ws[f'N{fin_row}'].number_format = currency_format
        ws[f'N{fin_row}'].border = thin_border
        ws[f'N{fin_row}'].fill = fin_template_fills['N']

        ws[f'O{fin_row}'] = record.occupied_rent
        ws[f'O{fin_row}'].number_format = currency_format
        ws[f'O{fin_row}'].border = thin_border
        ws[f'O{fin_row}'].fill = fin_template_fills['O']

        ws[f'P{fin_row}'] = record.revenue
        ws[f'P{fin_row}'].number_format = currency_format
        ws[f'P{fin_row}'].border = thin_border
        ws[f'P{fin_row}'].fill = fin_template_fills['P']

        ws[f'Q{fin_row}'] = record.expenses
        ws[f'Q{fin_row}'].number_format = currency_format
        ws[f'Q{fin_row}'].border = thin_border
        ws[f'Q{fin_row}'].fill = fin_template_fills['Q']

        ws[f'R{fin_row}'] = record.owed
        ws[f'R{fin_row}'].number_format = currency_format
        ws[f'R{fin_row}'].border = thin_border
        ws[f'R{fin_row}'].fill = fin_template_fills['R']

        ws[f'S{fin_row}'] = record.charges
        ws[f'S{fin_row}'].number_format = currency_format
        ws[f'S{fin_row}'].border = thin_border
        ws[f'S{fin_row}'].fill = fin_template_fills['S']

        # collections_pct stores as percentage (e.g., 92.32 = 92.32%)
        # Excel needs decimal (0.9232) with '0.00%' format
        if record.collections_pct:
            ws[f'T{fin_row}'] = record.collections_pct / 100
        elif record.collections:
            # Fallback to old collections column
            if record.collections > 1:
                ws[f'T{fin_row}'] = record.collections / 100
            else:
                ws[f'T{fin_row}'] = record.collections
        else:
            ws[f'T{fin_row}'] = None
        ws[f'T{fin_row}'].number_format = '0.00%'
        ws[f'T{fin_row}'].border = thin_border
        ws[f'T{fin_row}'].fill = fin_template_fills['T']

        fin_row += 1

    print(f"  Populated {len(historical_financial)} historical financial rows with borders and fill")

    # Set chart y-axis to auto-scale and adjust data ranges
    from openpyxl.chart.data_source import NumRef, NumDataSource
    from openpyxl.chart.label import DataLabelList

    # Find actual data ranges for each column
    def find_data_range(worksheet, col, start_row=3, max_row=200):
        """Find first and last non-empty row in a column."""
        first = None
        last = None
        for row in range(start_row, max_row):
            val = worksheet[f'{col}{row}'].value
            if val is not None:
                if first is None:
                    first = row
                last = row
        return first, last

    # Get data ranges for key columns (last 13 months only for charts)
    months_to_show = 13
    fin_end_row = 2 + len(historical_financial)
    if len(historical_financial) > months_to_show:
        fin_chart_start = fin_end_row - months_to_show + 1
    else:
        fin_chart_start = 3

    # Update charts with last 13 months of data
    for chart in ws._charts:
        # Check if this is the collections chart (uses column T)
        is_collections_chart = False
        for series in chart.series:
            if series.val and series.val.numRef and '$T$' in series.val.numRef.f:
                is_collections_chart = True
                break

        if is_collections_chart:
            # Set y-axis to 0-100% for collections chart (values are decimals 0-1)
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1.0
            chart.y_axis.majorUnit = 0.25
            chart.y_axis.number_format = '0%'
            chart.y_axis.delete = False
            chart.y_axis.tickLblPos = 'low'
        else:
            chart.y_axis.scaling.min = None
            chart.y_axis.scaling.max = None

        # Update each series to use last 13 months
        for series in chart.series:
            if series.val and series.val.numRef:
                ref = series.val.numRef.f
                # Determine which column this series uses and update range
                if '$N$' in ref:
                    new_ref = f"Financial!$N${fin_chart_start}:$N${fin_end_row}"
                    cat_ref = f"Financial!$M${fin_chart_start}:$M${fin_end_row}"
                    series.val.numRef.f = new_ref
                    if series.cat and series.cat.numRef:
                        series.cat.numRef.f = cat_ref
                elif '$O$' in ref:
                    new_ref = f"Financial!$O${fin_chart_start}:$O${fin_end_row}"
                    cat_ref = f"Financial!$M${fin_chart_start}:$M${fin_end_row}"
                    series.val.numRef.f = new_ref
                    if series.cat and series.cat.numRef:
                        series.cat.numRef.f = cat_ref
                elif '$P$' in ref:
                    new_ref = f"Financial!$P${fin_chart_start}:$P${fin_end_row}"
                    cat_ref = f"Financial!$M${fin_chart_start}:$M${fin_end_row}"
                    series.val.numRef.f = new_ref
                    if series.cat and series.cat.numRef:
                        series.cat.numRef.f = cat_ref
                elif '$Q$' in ref:
                    new_ref = f"Financial!$Q${fin_chart_start}:$Q${fin_end_row}"
                    cat_ref = f"Financial!$M${fin_chart_start}:$M${fin_end_row}"
                    series.val.numRef.f = new_ref
                    if series.cat and series.cat.numRef:
                        series.cat.numRef.f = cat_ref
                elif '$T$' in ref:
                    new_ref = f"Financial!$T${fin_chart_start}:$T${fin_end_row}"
                    cat_ref = f"Financial!$M${fin_chart_start}:$M${fin_end_row}"
                    series.val.numRef.f = new_ref
                    if series.cat and series.cat.numRef:
                        series.cat.numRef.f = cat_ref

    for chart in ws._charts:
        for series in chart.series:
            if series.val and series.val.numRef:
                ref = series.val.numRef.f
                # Show data labels for rent (N, O) and collections (T) charts
                if '$N$' in ref:
                    # Market Rent - labels above
                    labels = DataLabelList()
                    labels.showVal = True
                    labels.showPercent = False
                    labels.showCatName = False
                    labels.showSerName = False
                    labels.showLegendKey = False
                    labels.dLblPos = 't'  # top/above
                    series.labels = labels
                elif '$O$' in ref:
                    # Occupied Rent - labels below
                    labels = DataLabelList()
                    labels.showVal = True
                    labels.showPercent = False
                    labels.showCatName = False
                    labels.showSerName = False
                    labels.showLegendKey = False
                    labels.dLblPos = 'b'  # bottom/below
                    series.labels = labels
                elif '$T$' in ref:
                    # Collections - labels above, formatted as percentage
                    labels = DataLabelList()
                    labels.showVal = True
                    labels.showPercent = False
                    labels.showCatName = False
                    labels.showSerName = False
                    labels.showLegendKey = False
                    labels.dLblPos = 't'  # top/above
                    labels.numFmt = '0.00%'
                    series.labels = labels
                else:
                    # Remove data labels for other charts
                    labels = DataLabelList()
                    labels.showVal = False
                    labels.showPercent = False
                    labels.showCatName = False
                    labels.showSerName = False
                    labels.showLegendKey = False
                    series.labels = labels

    print(f"  Updated {len(ws._charts)} chart(s) with data labels on rent and collections")

    # Save
    wb.save(output_path)

    # Fix openpyxl relationship paths (it writes absolute paths, Excel needs relative)
    _fix_xlsx_rels(output_path, template_path)

    if session:
        session.close()

    print(f"\n✓ Report saved to: {output_path}")

    # Print summary
    print(f"\n{'='*50}")
    print(f"REPORT SUMMARY: {property_info['name']}")
    print(f"{'='*50}")
    print(f"  Property Code: {property_code}")
    print(f"  Report Date: {report_date}")
    print(f"  Units: {box_score.get('units', 0)}")
    print(f"  Occupancy: {box_score.get('pct_occupied', 0)*100:.2f}%")
    print(f"  Leased: {box_score.get('pct_leased', 0)*100:.2f}%")
    print(f"  Available: {box_score.get('available', 0)}")
    print(f"  Delinquency: ${delinquency.get('total_owed', 0):,.2f}")
    print(f"  Evictions Filed: {notices.get('evictions_filed', 0)}")
    print(f"  MTM Leases: {lease_exp.get('mtm', 0)}")
    print(f"  Make Readies: {make_ready_count}")
    print(f"  Work Orders: {work_order_count}")
    print(f"  Historical Occupancy Records: {len(historical_occupancy)}")
    print(f"  Historical Financial Records: {len(historical_financial)}")
    if property_info.get('equity_investment'):
        print(f"  Equity Investment: ${property_info['equity_investment']:,.2f}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_weekly_report.py <data_directory> [output_path] [report_date]")
        print("Example: python generate_weekly_report.py '/path/to/Perry Heights' ./report.xlsx 2026-02-09")
        sys.exit(1)

    data_dir = sys.argv[1]
    template_path = "/Users/shivaanikomanduri/Arcan_Weekly_Reports/templates/Weekly_Report_Template_Clean.xlsx"

    # Generate output filename from directory name
    property_name = os.path.basename(data_dir).replace(" ", "_")
    default_output = f"/Users/shivaanikomanduri/Arcan_Weekly_Reports/{property_name}_Weekly_Report.xlsx"
    output_path = sys.argv[2] if len(sys.argv) > 2 else default_output

    # Optional report date
    report_date = sys.argv[3] if len(sys.argv) > 3 else None

    generate_report(data_dir, template_path, output_path, report_date)
