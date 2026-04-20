"""Parse a long template weekly report INPUT sheet and return historical data as dicts."""
from openpyxl import load_workbook
from datetime import datetime


def _iter_dated_rows(ws, date_col, start_row, max_row=300):
    """Yield (date_str, row_idx) for rows where date_col is a datetime."""
    for r in range(start_row, max_row + 1):
        v = ws[f'{date_col}{r}'].value
        if isinstance(v, datetime):
            yield v.strftime('%Y-%m-%d'), r
        elif v is None:
            continue  # tolerate gaps


def _cell(ws, col, r):
    return ws[f'{col}{r}'].value


def parse_input(path):
    """Return (occupancy_rows, financial_rows) where each is a list of dicts.

    occupancy_rows: dicts with keys date, occupancy, turnover, work_orders, make_ready
    financial_rows: dicts with keys date, market_rent, occupied_rent, income_actual,
        income_budget, expense_actual, expense_budget, actual_cash, adjusted_cash,
        charges, collected, collections_pct
    All rows are merged by date per kind (monthly dates merge together, weekly dates separate).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['INPUT']

    occ = {}   # date -> dict
    fin = {}   # date -> dict

    def occ_set(d, **kv):
        occ.setdefault(d, {'date': d})
        for k, v in kv.items():
            if v is not None:
                occ[d][k] = v

    def fin_set(d, **kv):
        fin.setdefault(d, {'date': d})
        for k, v in kv.items():
            if v is not None:
                fin[d][k] = v

    # Historical Occupancy: M=date, N=occupancy%  (decimal → percentage)
    for d, r in _iter_dated_rows(ws, 'M', 2):
        val = _cell(ws, 'N', r)
        if val is not None:
            occ_set(d, occupancy=round(val * 100, 2))

    # Historical Turnover: P=date, Q=turnover% (decimal)
    for d, r in _iter_dated_rows(ws, 'P', 2):
        val = _cell(ws, 'Q', r)
        if val is not None:
            occ_set(d, turnover=round(val * 100, 2))

    # Work orders: AT=date, AU=WO count, AV=MR count (weekly)
    for d, r in _iter_dated_rows(ws, 'AT', 3):
        occ_set(d, work_orders=_cell(ws, 'AU', r), make_ready=_cell(ws, 'AV', r))

    # Income/Expense: S=date, T=inc_actual, U=inc_budget, W=exp_actual, X=exp_budget
    for d, r in _iter_dated_rows(ws, 'S', 3):
        fin_set(d,
                income_actual=_cell(ws, 'T', r),
                income_budget=_cell(ws, 'U', r),
                expense_actual=_cell(ws, 'W', r),
                expense_budget=_cell(ws, 'X', r))

    # Cash Balance: Z=date, AA=actual, AB=adjusted (starts row 5; row 3 is current week)
    for d, r in _iter_dated_rows(ws, 'Z', 3):
        actual = _cell(ws, 'AA', r)
        adjusted = _cell(ws, 'AB', r)
        # Skip header row where AA="ACTUAL "
        if isinstance(actual, str) or isinstance(adjusted, str):
            continue
        fin_set(d, actual_cash=actual, adjusted_cash=adjusted)

    # Collections: AD=date, AE=charges, AF=collected, AG=% collected (decimal)
    for d, r in _iter_dated_rows(ws, 'AD', 3):
        pct = _cell(ws, 'AG', r)
        fin_set(d,
                charges=_cell(ws, 'AE', r),
                collected=_cell(ws, 'AF', r),
                collections_pct=round(pct * 100, 2) if pct is not None else None)

    # Historical Rents: AP=date, AQ=market, AR=in-place
    for d, r in _iter_dated_rows(ws, 'AP', 3):
        fin_set(d,
                market_rent=_cell(ws, 'AQ', r),
                occupied_rent=_cell(ws, 'AR', r))

    # Drop rows that ended up with only the date key
    occ_rows = [v for v in occ.values() if len(v) > 1]
    fin_rows = [v for v in fin.values() if len(v) > 1]

    # Round financial floats to 2 decimals for storage cleanliness
    for row in fin_rows:
        for k, v in list(row.items()):
            if k != 'date' and isinstance(v, float):
                row[k] = round(v, 2)

    return occ_rows, fin_rows


if __name__ == '__main__':
    import sys, json
    occ, fin = parse_input(sys.argv[1])
    print(f"Occupancy rows: {len(occ)}")
    print(f"Financial rows: {len(fin)}")
    print("Sample occ:", json.dumps(occ[:2], default=str))
    print("Sample fin:", json.dumps(fin[:2], default=str))
