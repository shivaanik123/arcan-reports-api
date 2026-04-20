#!/usr/bin/env python3
"""
Long Report Generator
Generates weekly property reports for external ownership group properties.
Uses the Long_Report_Template.xlsx template with INPUT sheet structure.
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date, timedelta
import os
import sys
import glob
import re
import zipfile
import shutil
import tempfile
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker


def _fix_xlsx_rels(filepath, template_path=None):
    """
    Fix openpyxl xlsx output by copying .rels files from the original template.
    openpyxl rewrites relationship files with absolute paths and phantom references
    that Excel can't handle. The safest fix is to restore the original .rels files.
    """
    if template_path is None:
        # Try to find template
        for candidate in [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'Long_Report_Template.xlsx'),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'Weekly_Report_Template_Clean.xlsx'),
        ]:
            if os.path.exists(candidate):
                template_path = candidate
                break
    if not template_path or not os.path.exists(template_path):
        return  # Can't fix without template

    temp_path = filepath + '.tmp'
    with zipfile.ZipFile(template_path, 'r') as zt, \
         zipfile.ZipFile(filepath, 'r') as zin, \
         zipfile.ZipFile(temp_path, 'w') as zout:

        # Get template rels files
        template_rels = {n: zt.read(n) for n in zt.namelist() if n.endswith('.rels')}

        for item in zin.infolist():
            if item.filename.endswith('.rels') and item.filename in template_rels:
                # Use the template's version of this .rels file
                zout.writestr(item, template_rels[item.filename])
            else:
                zout.writestr(item, zin.read(item.filename))

    shutil.move(temp_path, filepath)

# Database connection
DATABASE_URL = "postgresql://postgres:ysBxQxKGOxlvIFfhCVrEYlApmrqElAMB@shinkansen.proxy.rlwy.net:55881/railway"

# Properties that use the long template
LONG_TEMPLATE_PROPERTIES = [
    "manwes", "tapeprk", "haven", "hampec", "talloak",
    "colwds", "marshp", "capella2", "55pharr", "emersn"
]


def get_db_session():
    """Get database session."""
    engine = create_engine(DATABASE_URL, echo=False)
    Session = sessionmaker(bind=engine)
    return Session()


def get_property_info(session, property_code):
    """Get property info from database."""
    result = session.execute(
        text("SELECT name, location, units, models FROM properties WHERE property_code = :code"),
        {"code": property_code}
    )
    row = result.fetchone()
    if row:
        return {
            'name': row[0],
            'location': row[1],
            'units': row[2],
            'models': row[3]
        }
    return None


def get_historical_occupancy(session, property_code, limit=53):
    """Get historical occupancy data (monthly records with occupancy %)."""
    result = session.execute(
        text("""
            SELECT date, occupancy FROM historical_occupancy
            WHERE property_code = :code AND occupancy IS NOT NULL
            ORDER BY date ASC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    return [{"date": row[0], "occupancy": row[1]} for row in result.fetchall()]


def get_historical_turnover(session, property_code, limit=53):
    """Get historical turnover data (monthly records with turnover %)."""
    result = session.execute(
        text("""
            SELECT date, turnover FROM historical_occupancy
            WHERE property_code = :code AND turnover IS NOT NULL
            ORDER BY date ASC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    return [{"date": row[0], "turnover": row[1]} for row in result.fetchall()]


def get_historical_income(session, property_code, limit=53):
    """Get historical income data (monthly records with actual/budget)."""
    result = session.execute(
        text("""
            SELECT date, income_actual, income_budget FROM historical_financial
            WHERE property_code = :code AND income_actual IS NOT NULL
            ORDER BY date ASC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    return [{"date": row[0], "actual": row[1], "budget": row[2]} for row in result.fetchall()]


def get_historical_expense(session, property_code, limit=53):
    """Get historical expense data (monthly records with actual/budget)."""
    result = session.execute(
        text("""
            SELECT date, expense_actual, expense_budget FROM historical_financial
            WHERE property_code = :code AND expense_actual IS NOT NULL
            ORDER BY date ASC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    return [{"date": row[0], "actual": row[1], "budget": row[2]} for row in result.fetchall()]


def get_historical_cash(session, property_code, limit=53):
    """Get historical cash balance data."""
    result = session.execute(
        text("""
            SELECT date, actual_cash, adjusted_cash FROM historical_financial
            WHERE property_code = :code AND actual_cash IS NOT NULL
            ORDER BY date ASC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    return [{"date": row[0], "actual": row[1], "adjusted": row[2]} for row in result.fetchall()]


def get_historical_collections(session, property_code, limit=53):
    """Get historical collections data."""
    result = session.execute(
        text("""
            SELECT date, charges, collected, collections_pct FROM historical_financial
            WHERE property_code = :code AND charges IS NOT NULL
            ORDER BY date ASC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    return [{"date": row[0], "charges": row[1], "collected": row[2], "pct": row[3]} for row in result.fetchall()]


def get_historical_rents(session, property_code, limit=53):
    """Get historical market and in-place rent data."""
    result = session.execute(
        text("""
            SELECT date, market_rent, occupied_rent FROM historical_financial
            WHERE property_code = :code AND market_rent IS NOT NULL
            ORDER BY date ASC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    return [{"date": row[0], "market": row[1], "in_place": row[2]} for row in result.fetchall()]


def get_historical_work_orders(session, property_code, limit=150):
    """Get historical work orders and make readies data (weekly records)."""
    result = session.execute(
        text("""
            SELECT date, work_orders, make_ready FROM historical_occupancy
            WHERE property_code = :code AND work_orders IS NOT NULL
            ORDER BY date ASC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    return [{"date": row[0], "work_orders": row[1], "make_ready": row[2]} for row in result.fetchall()]


def get_historical_budget(session, property_code, limit=3):
    """Get historical budget data (monthly records) for FINANCIAL section.

    Returns the most recent 3 months of budget data, ordered newest first.
    """
    result = session.execute(
        text("""
            SELECT date,
                   net_rental_income_actual, net_rental_income_budget,
                   other_income_actual, other_income_budget,
                   total_income_actual, total_income_budget,
                   payroll_benefits_actual, payroll_benefits_budget,
                   management_fee_actual, management_fee_budget,
                   general_admin_actual, general_admin_budget,
                   utilities_actual, utilities_budget,
                   repairs_maintenance_actual, repairs_maintenance_budget,
                   contract_services_actual, contract_services_budget,
                   make_ready_actual, make_ready_budget,
                   recreation_amenities_actual, recreation_amenities_budget,
                   advertising_marketing_actual, advertising_marketing_budget,
                   taxes_insurance_actual, taxes_insurance_budget,
                   total_operating_expenses_actual, total_operating_expenses_budget,
                   net_operating_income_actual, net_operating_income_budget,
                   debt_service_actual, debt_service_budget,
                   routine_replacement_actual, routine_replacement_budget,
                   capital_improvements_actual, capital_improvements_budget,
                   net_income_for_tax_actual, net_income_for_tax_budget
            FROM historical_budget
            WHERE property_code = :code
            ORDER BY date DESC LIMIT :limit
        """),
        {"code": property_code, "limit": limit}
    )
    rows = result.fetchall()
    budget_data = []
    for row in rows:
        budget_data.append({
            'date': row[0],
            'net_rental_income_actual': row[1], 'net_rental_income_budget': row[2],
            'other_income_actual': row[3], 'other_income_budget': row[4],
            'total_income_actual': row[5], 'total_income_budget': row[6],
            'payroll_benefits_actual': row[7], 'payroll_benefits_budget': row[8],
            'management_fee_actual': row[9], 'management_fee_budget': row[10],
            'general_admin_actual': row[11], 'general_admin_budget': row[12],
            'utilities_actual': row[13], 'utilities_budget': row[14],
            'repairs_maintenance_actual': row[15], 'repairs_maintenance_budget': row[16],
            'contract_services_actual': row[17], 'contract_services_budget': row[18],
            'make_ready_actual': row[19], 'make_ready_budget': row[20],
            'recreation_amenities_actual': row[21], 'recreation_amenities_budget': row[22],
            'advertising_marketing_actual': row[23], 'advertising_marketing_budget': row[24],
            'taxes_insurance_actual': row[25], 'taxes_insurance_budget': row[26],
            'total_operating_expenses_actual': row[27], 'total_operating_expenses_budget': row[28],
            'net_operating_income_actual': row[29], 'net_operating_income_budget': row[30],
            'debt_service_actual': row[31], 'debt_service_budget': row[32],
            'routine_replacement_actual': row[33], 'routine_replacement_budget': row[34],
            'capital_improvements_actual': row[35], 'capital_improvements_budget': row[36],
            'net_income_for_tax_actual': row[37], 'net_income_for_tax_budget': row[38],
        })
    return budget_data


def extract_budget_comparison_data(filepath):
    """Extract financial data from Budget Comparison report.

    Returns dict with actual and budget values for each financial category.
    """
    df = pd.read_excel(filepath)

    data = {
        'net_rental_income_actual': None, 'net_rental_income_budget': None,
        'other_income_actual': None, 'other_income_budget': None,
        'total_income_actual': None, 'total_income_budget': None,
        'payroll_benefits_actual': None, 'payroll_benefits_budget': None,
        'management_fee_actual': None, 'management_fee_budget': None,
        'general_admin_actual': None, 'general_admin_budget': None,
        'utilities_actual': None, 'utilities_budget': None,
        'repairs_maintenance_actual': None, 'repairs_maintenance_budget': None,
        'contract_services_actual': None, 'contract_services_budget': None,
        'make_ready_actual': None, 'make_ready_budget': None,
        'recreation_amenities_actual': None, 'recreation_amenities_budget': None,
        'advertising_marketing_actual': None, 'advertising_marketing_budget': None,
        'taxes_insurance_actual': None, 'taxes_insurance_budget': None,
        'total_operating_expenses_actual': None, 'total_operating_expenses_budget': None,
        'net_operating_income_actual': None, 'net_operating_income_budget': None,
        'debt_service_actual': None, 'debt_service_budget': None,
        'routine_replacement_actual': None, 'routine_replacement_budget': None,
        'capital_improvements_actual': None, 'capital_improvements_budget': None,
        'net_income_for_tax_actual': None, 'net_income_for_tax_budget': None,
    }

    # Search for each category by name in the data
    for idx, row in df.iterrows():
        name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        mtd_actual = row.iloc[2] if pd.notna(row.iloc[2]) else 0
        mtd_budget = row.iloc[3] if pd.notna(row.iloc[3]) else 0

        # Income categories
        if 'Total Rental Income' in name:
            data['net_rental_income_actual'] = mtd_actual
            data['net_rental_income_budget'] = mtd_budget
        elif 'Total Other Income' in name:
            data['other_income_actual'] = mtd_actual
            data['other_income_budget'] = mtd_budget
        elif name == 'TOTAL INCOME':
            data['total_income_actual'] = mtd_actual
            data['total_income_budget'] = mtd_budget

        # Expense categories
        elif 'Total Payroll & Benefits' in name:
            data['payroll_benefits_actual'] = mtd_actual
            data['payroll_benefits_budget'] = mtd_budget
        elif 'Total Management Fees' in name:
            data['management_fee_actual'] = mtd_actual
            data['management_fee_budget'] = mtd_budget
        elif 'Total General & Admin' in name:
            data['general_admin_actual'] = mtd_actual
            data['general_admin_budget'] = mtd_budget
        elif 'Total Utilities' in name:
            data['utilities_actual'] = mtd_actual
            data['utilities_budget'] = mtd_budget
        elif 'Total Repairs & Maintenance' in name:
            data['repairs_maintenance_actual'] = mtd_actual
            data['repairs_maintenance_budget'] = mtd_budget
        elif 'Total Contract Services' in name:
            data['contract_services_actual'] = mtd_actual
            data['contract_services_budget'] = mtd_budget
        elif 'Total Make Ready' in name:
            data['make_ready_actual'] = mtd_actual
            data['make_ready_budget'] = mtd_budget
        elif 'Total Recreation' in name or 'Recreation Amenities' in name:
            data['recreation_amenities_actual'] = mtd_actual
            data['recreation_amenities_budget'] = mtd_budget
        elif 'Total Adertising' in name or 'Total Advertising' in name:
            data['advertising_marketing_actual'] = mtd_actual
            data['advertising_marketing_budget'] = mtd_budget
        elif 'Total Taxes & Insurance' in name:
            data['taxes_insurance_actual'] = mtd_actual
            data['taxes_insurance_budget'] = mtd_budget
        elif name == 'TOTAL OPERATING EXPENSES':
            data['total_operating_expenses_actual'] = mtd_actual
            data['total_operating_expenses_budget'] = mtd_budget
        elif 'NET OPERATING INCOME' in name and 'AFTER' not in name:
            data['net_operating_income_actual'] = mtd_actual
            data['net_operating_income_budget'] = mtd_budget
        elif 'Total Debt Service' in name:
            data['debt_service_actual'] = mtd_actual
            data['debt_service_budget'] = mtd_budget
        elif 'Routine Replacements' in name and 'Subtotal' in name:
            data['routine_replacement_actual'] = mtd_actual
            data['routine_replacement_budget'] = mtd_budget
        elif 'Capital Improvements' in name and 'Subtotal' in name:
            data['capital_improvements_actual'] = mtd_actual
            data['capital_improvements_budget'] = mtd_budget
        elif 'TOTAL NET INCOME FOR TAX' in name:
            data['net_income_for_tax_actual'] = mtd_actual
            data['net_income_for_tax_budget'] = mtd_budget

    return data


def find_file(directory, pattern, property_code=None):
    """Find a file matching the pattern in the directory, excluding temp files.

    If property_code is provided, prefer files containing the property code.
    """
    matches = [f for f in glob.glob(os.path.join(directory, pattern))
               if not os.path.basename(f).startswith('~$')]

    if not matches:
        return None

    # If property_code provided, filter to files containing that code
    if property_code:
        code_matches = [f for f in matches if property_code.lower() in os.path.basename(f).lower()]
        if code_matches:
            return code_matches[0]

    # Return first match if no property-specific match found
    return matches[0]


def extract_box_score_data(filepath):
    """Extract occupancy and unit data from Box Score Summary."""
    df = pd.read_excel(filepath)

    data = {
        'units': 0,
        'occupied_no_notice': 0,
        'vacant_rented': 0,
        'vacant_unrented': 0,
        'notice_rented': 0,
        'notice_unrented': 0,
    }

    for idx, row in df.iterrows():
        if pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() == 'Total':
            data['units'] = int(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
            data['occupied_no_notice'] = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
            data['vacant_rented'] = int(row.iloc[6]) if pd.notna(row.iloc[6]) else 0
            data['vacant_unrented'] = int(row.iloc[7]) if pd.notna(row.iloc[7]) else 0
            data['notice_rented'] = int(row.iloc[8]) if pd.notna(row.iloc[8]) else 0
            data['notice_unrented'] = int(row.iloc[9]) if pd.notna(row.iloc[9]) else 0
            break

    return data


def shorten_unit_name(name):
    """Convert long unit names to short format.

    Examples:
        "1 bed 1 bath" -> "1x1"
        "1 bed 1 bath - C Building" -> "1x1 C Bldg"
        "2 bed 1 bath Upgrade" -> "2x1 Upgrade"
    """
    import re

    # Extract bed and bath counts
    match = re.match(r'(\d+)\s*bed\s*(\d+)\s*bath', name, re.IGNORECASE)
    if not match:
        return name  # Return original if pattern doesn't match

    beds = match.group(1)
    baths = match.group(2)
    short_name = f"{beds}x{baths}"

    # Get the suffix (everything after "X bed Y bath")
    suffix = name[match.end():].strip()

    # Shorten common suffixes
    if suffix:
        suffix = suffix.lstrip('- ')  # Remove leading dash/space
        suffix = suffix.replace('Building', 'Bldg')
        short_name = f"{short_name} {suffix}"

    return short_name


def extract_unit_mix_data(filepath):
    """Extract unit mix data (MARKET AND IN PLACE) from Box Score Summary.

    Returns list of dicts with unit type details:
    - unit: Unit name (e.g., "1x1", "1x1 C Bldg")
    - type: Unit type code (e.g., "1101a1")
    - units: Total number of units
    - occupied: Number of occupied units
    - sf: Square footage
    - market_rent: Average rent (market)
    - in_place_rent: In-place rent (same as market for now)
    """
    # Read with header row detection
    df = pd.read_excel(filepath, header=5)  # Header is typically on row 6

    unit_mix = []
    seen_types = set()  # Track unit types to avoid duplicates

    for idx, row in df.iterrows():
        name = row.get('Name')
        if name is None or pd.isna(name):
            continue

        name_str = str(name).strip()

        # Skip header, total, empty, NOT OWNED, HOA, and Not Specified rows
        if name_str == '':
            continue
        skip_keywords = ['not specified', 'not owned', 'hoa']
        name_lower = name_str.lower()
        if name_lower in ['name', 'total'] or any(kw in name_lower for kw in skip_keywords):
            continue

        # Get type code
        unit_type = str(row.get('Code', '')).strip() if pd.notna(row.get('Code')) else ''

        # Skip if we've already seen this unit type (avoid duplicates)
        if unit_type in seen_types:
            continue

        # Get unit count
        units = int(row.get('Units', 0)) if pd.notna(row.get('Units')) else 0

        # Skip rows with 0 units
        if units == 0:
            continue

        # Mark this type as seen
        seen_types.add(unit_type)

        # Get occupied count (Occupied No Notice + Notice Rented + Notice Unrented)
        occupied_no_notice = int(row.get('Occupied No Notice', 0)) if pd.notna(row.get('Occupied No Notice')) else 0
        notice_rented = int(row.get('Notice Rented', 0)) if pd.notna(row.get('Notice Rented')) else 0
        notice_unrented = int(row.get('Notice Unrented', 0)) if pd.notna(row.get('Notice Unrented')) else 0
        occupied = occupied_no_notice + notice_rented + notice_unrented

        # Get square footage
        def parse_sf(val):
            """Parse square footage, handling comma-formatted strings like '1,278'."""
            if pd.isna(val) or val == 0:
                return 0
            if isinstance(val, str):
                val = val.replace(',', '').strip()
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return 0

        sf = parse_sf(row.get('Avg. Sq Ft.', 0))
        # Try alternate column names
        if sf == 0:
            sf = parse_sf(row.get('Sqft', 0))
        if sf == 0:
            sf = parse_sf(row.get('SF', 0))
        if sf == 0:
            sf = parse_sf(row.get('Sq Ft', 0))

        # Get average rent (market rent)
        avg_rent_raw = row.get('Avg. Rent', 0)
        if pd.notna(avg_rent_raw):
            # Handle string format with commas
            if isinstance(avg_rent_raw, str):
                avg_rent_raw = avg_rent_raw.replace(',', '').replace('$', '')
            market_rent = float(avg_rent_raw) if avg_rent_raw else 0
        else:
            market_rent = 0

        unit_mix.append({
            'unit': shorten_unit_name(name_str),
            'type': unit_type,
            'units': units,
            'occupied': occupied,
            'sf': sf,
            'market_rent': market_rent,
            'in_place_rent': market_rent,  # Use same value for now
        })

    return unit_mix


def extract_conversion_data(filepath):
    """Extract traffic and application data from Conversion Ratios."""
    df = pd.read_excel(filepath)

    data = {
        'call': 0,
        'walkin': 0,
        'email': 0,
        'other': 0,
        'sms': 0,
        'web': 0,
        'chat': 0,
        'unqualified': 0,
        'total_traffic': 0,
        'applications': 0,
        'approved': 0,
        'denied': 0,
        'cancels': 0
    }

    for idx, row in df.iterrows():
        row_str = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        # Look for property code row (has numeric data)
        if row_str and row_str not in ['Property', ''] and 'Total' not in str(row.iloc[1] if pd.notna(row.iloc[1]) else ''):
            if pd.notna(row.iloc[3]):  # Has Call data
                data['call'] = int(row.iloc[3]) if pd.notna(row.iloc[3]) else 0
                data['walkin'] = int(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
                data['email'] = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
                data['other'] = int(row.iloc[6]) if pd.notna(row.iloc[6]) else 0
                data['sms'] = int(row.iloc[7]) if pd.notna(row.iloc[7]) else 0
                data['web'] = int(row.iloc[8]) if pd.notna(row.iloc[8]) else 0
                data['chat'] = int(row.iloc[9]) if pd.notna(row.iloc[9]) else 0
                data['unqualified'] = int(row.iloc[10]) if pd.notna(row.iloc[10]) else 0
                data['applications'] = int(row.iloc[12]) if pd.notna(row.iloc[12]) else 0
                data['approved'] = int(row.iloc[13]) if pd.notna(row.iloc[13]) else 0
                data['denied'] = int(row.iloc[17]) if pd.notna(row.iloc[17]) else 0
                data['cancels'] = int(row.iloc[18]) if pd.notna(row.iloc[18]) else 0
                break

    # Calculate total traffic
    data['total_traffic'] = (data['call'] + data['walkin'] + data['email'] +
                              data['other'] + data['sms'] + data['web'] +
                              data['chat'] + data['unqualified'])

    return data


def extract_residents_on_notice(filepath):
    """Extract notice and eviction counts from Residents on Notice file."""
    df = pd.read_excel(filepath)

    notice_count = 0
    eviction_count = 0

    for idx, row in df.iterrows():
        status = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''
        if status == 'Notice':
            notice_count += 1
        elif status == 'Eviction':
            eviction_count += 1

    return {
        'notice_count': notice_count,
        'eviction_count': eviction_count
    }


def extract_projected_occupancy(filepath):
    """Extract 6 weeks of projected move ins/outs from Projected Occupancy report.

    Returns list of dicts with date (end of week), move_ins, move_outs for 6 weeks.
    """
    df = pd.read_excel(filepath)

    projections = []

    for idx, row in df.iterrows():
        date_val = row.iloc[0]
        # Look for date patterns (rows with date in first column)
        if pd.notna(date_val):
            # Check if it's a date string like "02/09/2026" or datetime
            if isinstance(date_val, datetime):
                week_date = date_val.date()
            elif isinstance(date_val, str) and '/' in date_val and len(date_val) == 10:
                try:
                    week_date = datetime.strptime(date_val, '%m/%d/%Y').date()
                except ValueError:
                    continue
            else:
                continue

            move_ins = int(row.iloc[2]) if pd.notna(row.iloc[2]) else 0
            move_outs = int(row.iloc[3]) if pd.notna(row.iloc[3]) else 0

            projections.append({
                'date': week_date,
                'move_ins': move_ins,
                'move_outs': move_outs
            })

            # Stop after 6 weeks
            if len(projections) >= 6:
                break

    return projections


def get_report_sunday(report_date=None):
    """Get the Sunday of the current week for report date."""
    if report_date is None:
        today = date.today()
    elif isinstance(report_date, str):
        today = datetime.strptime(report_date, '%Y-%m-%d').date()
    else:
        today = report_date

    # Calculate Sunday of the current week
    # weekday() returns 0 for Monday, 6 for Sunday
    days_since_sunday = (today.weekday() + 1) % 7
    sunday = today - timedelta(days=days_since_sunday)

    return sunday


def copy_drawings_from_template(template_path, output_path):
    """Copy drawing XML files from template to output to preserve text boxes/shapes."""
    import zipfile
    import shutil
    import tempfile
    import os

    # Create a temporary copy of the output
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp_path = tmp.name

    shutil.copy(output_path, tmp_path)

    # Open both files and copy drawings from template
    with zipfile.ZipFile(template_path, 'r') as template_zip:
        with zipfile.ZipFile(tmp_path, 'r') as output_zip:
            with zipfile.ZipFile(output_path, 'w') as new_zip:
                # Copy all files from output except drawings
                for item in output_zip.namelist():
                    if 'drawing' not in item.lower():
                        new_zip.writestr(item, output_zip.read(item))

                # Copy drawings from template
                for item in template_zip.namelist():
                    if 'drawing' in item.lower():
                        new_zip.writestr(item, template_zip.read(item))

    os.unlink(tmp_path)


def generate_long_report(data_dir, template_path, output_path, property_code, report_date=None):
    """Generate a long format weekly report from source data and database."""

    print(f"Generating long report for: {property_code}")
    print(f"Data directory: {data_dir}")

    # Get database session
    session = get_db_session()

    # Get property info from database
    property_info = get_property_info(session, property_code)
    if property_info:
        print(f"  Property name: {property_info['name']}")
        print(f"  Location: {property_info['location']}")
        print(f"  Units: {property_info['units']}")
        print(f"  Model/Down: {property_info['models']}")
    else:
        print(f"  Warning: No property info found in database for '{property_code}'")
        property_info = {'name': property_code, 'location': '', 'units': 0, 'models': '0'}

    # Find source files (prefer files matching property code)
    box_score_file = find_file(data_dir, "*Box_Score*.xlsx", property_code)
    conversion_file = find_file(data_dir, "*Conversion_Ratios*.xlsx", property_code)
    notice_file = find_file(data_dir, "*Residents_on_Notice*.xlsx", property_code)
    projected_occ_file = find_file(data_dir, "*Projected*Occupancy*.xlsx", property_code)
    if not projected_occ_file:
        projected_occ_file = find_file(data_dir, "*ProjectedOccupancy*.xlsx", property_code)
    # Find all budget comparison files for the property (up to 3 months)
    budget_comparison_files = [f for f in glob.glob(os.path.join(data_dir, "*Budget_Comparison*.xlsx"))
                               if not os.path.basename(f).startswith('~$') and
                               property_code.lower() in os.path.basename(f).lower()]
    # Sort by filename (assumes month info is in filename) - most recent first
    budget_comparison_files.sort(reverse=True)
    budget_comparison_files = budget_comparison_files[:3]  # Take up to 3 files

    print(f"\n  Source files found:")
    print(f"    Box Score: {os.path.basename(box_score_file) if box_score_file else 'NOT FOUND'}")
    print(f"    Conversion: {os.path.basename(conversion_file) if conversion_file else 'NOT FOUND'}")
    print(f"    Notice: {os.path.basename(notice_file) if notice_file else 'NOT FOUND'}")
    print(f"    Projected Occupancy: {os.path.basename(projected_occ_file) if projected_occ_file else 'NOT FOUND'}")
    print(f"    Budget Comparison: {len(budget_comparison_files)} file(s)")
    for bc_file in budget_comparison_files:
        print(f"      - {os.path.basename(bc_file)}")

    # Extract data
    print("\n  Extracting data from source files...")
    box_score = extract_box_score_data(box_score_file) if box_score_file else {}
    unit_mix = extract_unit_mix_data(box_score_file) if box_score_file else []
    conversion = extract_conversion_data(conversion_file) if conversion_file else {}
    notices = extract_residents_on_notice(notice_file) if notice_file else {}
    projected_occ = extract_projected_occupancy(projected_occ_file) if projected_occ_file else []
    # Extract data from up to 3 budget comparison files
    budget_data_list = []
    for bc_file in budget_comparison_files:
        budget_data_list.append(extract_budget_comparison_data(bc_file))

    # Print unit mix data
    if unit_mix:
        print(f"\n  Unit Mix ({len(unit_mix)} unit types):")
        for um in unit_mix:
            print(f"    {um['unit']} ({um['type']}): {um['units']} units, {um['occupied']} occ, {um['sf']} SF")

    # Calculate report dates
    sunday_date = get_report_sunday(report_date)
    week_end_date = sunday_date + timedelta(days=7)

    print(f"\n  Report dates:")
    print(f"    Sunday (C2): {sunday_date}")
    print(f"    Week End (C6): {week_end_date}")

    # Calculate derived values
    total_occupied = (box_score.get('occupied_no_notice', 0) +
                      box_score.get('notice_rented', 0) +
                      box_score.get('notice_unrented', 0))
    vacant_rentable = (box_score.get('vacant_rented', 0) +
                       box_score.get('vacant_unrented', 0))
    leased_vacant = box_score.get('vacant_rented', 0)

    print(f"\n  Calculated values:")
    print(f"    Total Occupied: {total_occupied}")
    print(f"    Vacant Rentable: {vacant_rentable}")
    print(f"    Leased Vacant: {leased_vacant}")
    print(f"    Notice Units: {notices.get('notice_count', 0)}")
    print(f"    Eviction Units: {notices.get('eviction_count', 0)}")
    print(f"    Total Traffic: {conversion.get('total_traffic', 0)}")
    print(f"    Applications: {conversion.get('applications', 0)}")

    # Load template
    print(f"\n  Loading template: {template_path}")
    wb = load_workbook(template_path)

    # Get INPUT sheet
    ws = wb['INPUT']

    # Populate INPUT sheet
    print("  Populating INPUT sheet...")

    # Section 1: Property Info
    ws['C2'] = sunday_date  # Date
    ws['C3'] = property_info['name']  # Property
    ws['C4'] = property_info['location']  # Location
    ws['C5'] = property_info['units']  # Units
    ws['C6'] = week_end_date  # Week End
    ws['C8'] = total_occupied  # Total Occupied Units
    ws['C9'] = int(property_info['models']) if property_info['models'] else 0  # Model/Down Units
    ws['C10'] = vacant_rentable  # Vacant Rentable Units
    ws['C11'] = leased_vacant  # Leased Vacant Units
    ws['C12'] = notices.get('notice_count', 0)  # Notice Units
    ws['C13'] = notices.get('eviction_count', 0)  # Units Under Eviction
    ws['C14'] = box_score.get('notice_rented', 0)  # Pre-Leased Notice/Eviction

    # Section 2: Leasing Activity
    ws['C16'] = conversion.get('total_traffic', 0)  # Total Traffic
    ws['C17'] = conversion.get('applications', 0)  # Total Applications
    ws['C18'] = conversion.get('approved', 0)  # Applications Approved
    ws['C19'] = conversion.get('cancels', 0)  # Applications Cancelled
    ws['C20'] = conversion.get('denied', 0)  # Applications Denied

    # Section: FINANCIAL (Columns E-K) - Pull from historical_budget database table
    # E = Category labels
    # F/G = Month 1 Actual/Budget, H/I = Month 2 Actual/Budget, J/K = Month 3 Actual/Budget
    hist_budget = get_historical_budget(session, property_code, limit=3)
    if hist_budget:
        print(f"  Populating FINANCIAL section ({len(hist_budget)} month(s) from database)...")
        for hb in hist_budget:
            print(f"    {hb['date'].strftime('%B %Y')}: Net Rental Income Actual=${hb['net_rental_income_actual']:,.2f}")

        # Column mapping for each month: (actual_col, budget_col)
        # Data is ordered newest first, so Month 1 = most recent
        month_columns = [('F', 'G'), ('H', 'I'), ('J', 'K')]

        # Row mapping for each category
        category_rows = {
            'net_rental_income': 3,
            'other_income': 4,
            'total_income': 5,
            'payroll_benefits': 8,
            'management_fee': 9,
            'general_admin': 10,
            'utilities': 11,
            'repairs_maintenance': 12,
            'contract_services': 13,
            'make_ready': 14,
            'recreation_amenities': 15,
            'advertising_marketing': 16,
            'taxes_insurance': 17,
            'total_operating_expenses': 18,
            'net_operating_income': 20,
            'debt_service': 22,
            'routine_replacement': 23,
            'capital_improvements': 24,
            'net_income_for_tax': 26,
        }

        for month_idx, budget_data in enumerate(hist_budget):
            if month_idx >= 3:
                break
            actual_col, budget_col = month_columns[month_idx]

            for category, row in category_rows.items():
                actual_val = budget_data.get(f'{category}_actual')
                budget_val = budget_data.get(f'{category}_budget')
                if actual_val is not None:
                    ws[f'{actual_col}{row}'] = actual_val
                if budget_val is not None:
                    ws[f'{budget_col}{row}'] = budget_val
    else:
        print("  FINANCIAL section: No data found in historical_budget table")

    # Section 3: Scheduled Move Ins/Outs (Rows 24-29, 6 weeks)
    # Calculate dates as 6 weeks out from week end date (C6)
    # Use move ins/outs numbers from Projected Occupancy report
    if projected_occ:
        print(f"  Populating 6 weeks of projected occupancy...")
        for i in range(6):
            row_num = 24 + i
            # Calculate end of week date: week_end + i*7 days
            future_date = week_end_date + timedelta(days=i * 7)
            ws[f'A{row_num}'] = future_date  # End of week date
            # Use move ins/outs from report if available
            if i < len(projected_occ):
                ws[f'B{row_num}'] = projected_occ[i]['move_ins']
                ws[f'C{row_num}'] = projected_occ[i]['move_outs']
            else:
                ws[f'B{row_num}'] = 0
                ws[f'C{row_num}'] = 0

    # ==========================================
    # HISTORICAL DATA SECTIONS FROM DATABASE
    # ==========================================
    print("\n  Loading historical data from database...")

    # Section: Historical Occupancy (Column M)
    # M1 = header, M2+ = Date, N2+ = Occupancy %
    hist_occ = get_historical_occupancy(session, property_code)
    if hist_occ:
        print(f"    Historical Occupancy: {len(hist_occ)} records")
        for i, record in enumerate(hist_occ):
            row_num = 2 + i
            ws[f'M{row_num}'] = record['date']
            ws[f'N{row_num}'] = record['occupancy'] / 100 if record['occupancy'] else None  # Convert to decimal

    # Section: Historical Turnover (Column P)
    # P1 = header, P2+ = Date, Q2+ = Turnover %
    hist_turn = get_historical_turnover(session, property_code)
    if hist_turn:
        print(f"    Historical Turnover: {len(hist_turn)} records")
        for i, record in enumerate(hist_turn):
            row_num = 2 + i
            ws[f'P{row_num}'] = record['date']
            # Always show turnover value (0% should display, not be blank)
            ws[f'Q{row_num}'] = record['turnover'] / 100 if record['turnover'] is not None else 0

    # Section: Historical Income (Column S)
    # S1 = header, S2 = "CURRENT" label + sub-headers (don't overwrite), S3 = column headers, data starts row 4
    hist_income = get_historical_income(session, property_code)
    if hist_income:
        print(f"    Historical Income: {len(hist_income)} records")
        # Historical data starts at row 4 (oldest first)
        for i, record in enumerate(hist_income):
            row_num = 4 + i
            ws[f'S{row_num}'] = record['date']
            ws[f'T{row_num}'] = record['actual']
            ws[f'U{row_num}'] = record['budget']

    # Section: Historical Expense (Column W)
    # W1 = header, W2 = "CURRENT" label + sub-headers (don't overwrite), W3 = column headers, data starts row 4
    hist_expense = get_historical_expense(session, property_code)
    if hist_expense:
        print(f"    Historical Expense: {len(hist_expense)} records")
        # Historical data starts at row 4 (oldest first)
        for i, record in enumerate(hist_expense):
            row_num = 4 + i
            ws[f'W{row_num}'] = record['date']
            ws[f'X{row_num}'] = record['actual']
            ws[f'Y{row_num}'] = record['budget']

    # Section: Cash Balance (Column Z)
    # Z1 = header, Z2 = "CURRENT" label + sub-headers (don't overwrite), Z3 = column headers, data starts row 4
    hist_cash = get_historical_cash(session, property_code)
    if hist_cash:
        print(f"    Historical Cash: {len(hist_cash)} records")
        # Historical data starts at row 4 (oldest first)
        for i, record in enumerate(hist_cash):
            row_num = 4 + i
            ws[f'Z{row_num}'] = record['date']
            ws[f'AA{row_num}'] = record['actual']
            ws[f'AB{row_num}'] = record['adjusted']

    # Section: Collections (Column AD-AG)
    # AD1 = header, AD2 = "Date", AE2 = "CHARGES", AF2 = "COLLECTED", AG2 = "% collected"
    # Data starts row 3
    hist_coll = get_historical_collections(session, property_code)
    if hist_coll:
        print(f"    Historical Collections: {len(hist_coll)} records")
        for i, record in enumerate(hist_coll):
            row_num = 3 + i
            ws[f'AD{row_num}'] = record['date']
            ws[f'AE{row_num}'] = record['charges']
            ws[f'AF{row_num}'] = record['collected']
            ws[f'AG{row_num}'] = record['pct'] / 100 if record['pct'] else None  # Convert to decimal for %

    # Section: Market and In-Place Rents (Column AQ-AS)
    # AQ1 = header, AQ2 = "Date", AR2 = "MARKET", AS2 = "IN-PLACE"
    # Data starts row 3
    hist_rents = get_historical_rents(session, property_code)
    if hist_rents:
        print(f"    Historical Rents: {len(hist_rents)} records")
        for i, record in enumerate(hist_rents):
            row_num = 3 + i
            ws[f'AQ{row_num}'] = record['date']
            ws[f'AR{row_num}'] = record['market']
            ws[f'AS{row_num}'] = record['in_place']

    # Section: Work Orders (Column AU-AW)
    # AU1 = header, AU2 = "Date", AV2 = "Work Orders", AW2 = "Make Readies"
    # Data starts row 3
    hist_wo = get_historical_work_orders(session, property_code)
    if hist_wo:
        print(f"    Historical Work Orders: {len(hist_wo)} records")
        for i, record in enumerate(hist_wo):
            row_num = 3 + i
            ws[f'AU{row_num}'] = record['date']
            ws[f'AV{row_num}'] = record['work_orders']
            ws[f'AW{row_num}'] = record['make_ready']

    # Section: Unit Mix / MARKET AND IN PLACE (Columns AI-AO)
    # AI2 = "UNIT", AJ2 = "TYPE", AK2 = "# Units", AL2 = "# Occ", AM2 = "SF", AN2 = "MARKET", AO2 = "IN-PLACE"
    # Data starts row 3
    if unit_mix:
        print(f"    Unit Mix: {len(unit_mix)} unit types")
        from openpyxl.styles import Alignment, Font
        currency_fmt = '"$" #,##0.00'
        center = Alignment(horizontal='center')
        blue = Font(color='FF0000FF')

        for i, um in enumerate(unit_mix):
            row_num = 3 + i
            ws[f'AI{row_num}'] = um['unit']
            ws[f'AJ{row_num}'] = um['type']
            ws[f'AK{row_num}'] = um['units']
            ws[f'AL{row_num}'] = um['occupied']
            ws[f'AM{row_num}'] = um['sf']
            ws[f'AN{row_num}'] = um['market_rent']
            ws[f'AO{row_num}'] = um['in_place_rent']

            # Apply formatting - blue font, center alignment
            for col in ['AI', 'AJ', 'AK', 'AL', 'AM']:
                ws[f'{col}{row_num}'].font = blue
                ws[f'{col}{row_num}'].alignment = center
            # Currency formatting for rent columns
            ws[f'AN{row_num}'].number_format = currency_fmt
            ws[f'AN{row_num}'].font = blue
            ws[f'AN{row_num}'].alignment = center
            ws[f'AO{row_num}'].number_format = currency_fmt
            ws[f'AO{row_num}'].font = blue
            ws[f'AO{row_num}'].alignment = center

    # Apply currency formatting, alignment, and blue font to dollar columns
    from openpyxl.styles import Alignment, Font
    print("  Applying formatting...")
    currency_format = '"$" #,##0.00'
    center_align = Alignment(horizontal='center')
    blue_font = Font(color='FF0000FF')

    # Format FINANCIAL section (columns F-K for 3 months, rows 3-26)
    financial_rows = [3, 4, 5, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 20, 22, 23, 24, 26]
    financial_cols = ['F', 'G', 'H', 'I', 'J', 'K']
    for row in financial_rows:
        for col in financial_cols:
            if ws[f'{col}{row}'].value is not None:
                ws[f'{col}{row}'].number_format = currency_format
                ws[f'{col}{row}'].alignment = center_align
                ws[f'{col}{row}'].font = blue_font

    # Format data rows (row 4+) - don't touch row 2 headers
    for row in list(range(4, 60)):
        # Income: Actual (T), Budget (U)
        if ws[f'T{row}'].value is not None:
            ws[f'T{row}'].number_format = currency_format
            ws[f'T{row}'].alignment = center_align
            ws[f'T{row}'].font = blue_font
        if ws[f'U{row}'].value is not None:
            ws[f'U{row}'].number_format = currency_format
            ws[f'U{row}'].alignment = center_align
            ws[f'U{row}'].font = blue_font
        # Expense: Actual (X), Budget (Y)
        if ws[f'X{row}'].value is not None:
            ws[f'X{row}'].number_format = currency_format
            ws[f'X{row}'].alignment = center_align
            ws[f'X{row}'].font = blue_font
        if ws[f'Y{row}'].value is not None:
            ws[f'Y{row}'].number_format = currency_format
            ws[f'Y{row}'].alignment = center_align
            ws[f'Y{row}'].font = blue_font
        # Cash: Actual (AA), Adjusted (AB)
        if ws[f'AA{row}'].value is not None:
            ws[f'AA{row}'].number_format = currency_format
            ws[f'AA{row}'].alignment = center_align
            ws[f'AA{row}'].font = blue_font
        if ws[f'AB{row}'].value is not None:
            ws[f'AB{row}'].number_format = currency_format
            ws[f'AB{row}'].alignment = center_align
            ws[f'AB{row}'].font = blue_font
        # Collections: Charges (AE), Collected (AF)
        if ws[f'AE{row}'].value is not None:
            ws[f'AE{row}'].number_format = currency_format
            ws[f'AE{row}'].alignment = center_align
            ws[f'AE{row}'].font = blue_font
        if ws[f'AF{row}'].value is not None:
            ws[f'AF{row}'].number_format = currency_format
            ws[f'AF{row}'].alignment = center_align
            ws[f'AF{row}'].font = blue_font
        # Market/In-Place Rents (AR, AS)
        if ws[f'AR{row}'].value is not None:
            ws[f'AR{row}'].number_format = currency_format
            ws[f'AR{row}'].alignment = center_align
            ws[f'AR{row}'].font = blue_font
        if ws[f'AS{row}'].value is not None:
            ws[f'AS{row}'].number_format = currency_format
            ws[f'AS{row}'].alignment = center_align
            ws[f'AS{row}'].font = blue_font

    # Save
    wb.save(output_path)
    session.close()

    # Fix openpyxl relationship paths (it writes absolute paths, Excel needs relative)
    _fix_xlsx_rels(output_path, template_path)

    print(f"\n✓ Report saved to: {output_path}")

    # Print summary
    print(f"\n{'='*50}")
    print(f"REPORT SUMMARY: {property_info['name']}")
    print(f"{'='*50}")
    print(f"  Property Code: {property_code}")
    print(f"  Report Date (Sunday): {sunday_date}")
    print(f"  Week End: {week_end_date}")
    print(f"  Units: {property_info['units']}")
    print(f"  Total Occupied: {total_occupied}")
    print(f"  Vacant Rentable: {vacant_rentable}")
    print(f"  Notice Units: {notices.get('notice_count', 0)}")
    print(f"  Evictions: {notices.get('eviction_count', 0)}")
    print(f"  Total Traffic: {conversion.get('total_traffic', 0)}")
    print(f"  Applications: {conversion.get('applications', 0)}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_long_report.py <data_directory> <property_code> [output_path] [report_date]")
        print("Example: python generate_long_report.py '/path/to/55 Pharr' 55pharr ./report.xlsx 2026-02-09")
        sys.exit(1)

    data_dir = sys.argv[1]
    property_code = sys.argv[2]
    template_path = "/Users/shivaanikomanduri/Arcan_Weekly_Reports/templates/Long_Report_Template.xlsx"

    # Generate output filename
    property_name = os.path.basename(data_dir).replace(" ", "_")
    default_output = f"/Users/shivaanikomanduri/Arcan_Weekly_Reports/{property_name}_Long_Report.xlsx"
    output_path = sys.argv[3] if len(sys.argv) > 3 else default_output

    # Optional report date
    report_date = sys.argv[4] if len(sys.argv) > 4 else None

    generate_long_report(data_dir, template_path, output_path, property_code, report_date)
